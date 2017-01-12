#!/usr/bin/env python

##Working and copying from the 'openpyxl' source page at...
#'https://openpyxl.readthedocs.io/en/default'


import pip
#'openpyxl' is the external module to install
def install(openpyxl):
    pip.main(['install', openpyxl])

from openpyxl import Workbook
wb=Workbook()

#grab the active worksheet
ws=wb.active

#Data can be assigned directly to the cells
ws['A1']=42

#Rows can also be appended
ws.append([1, 2, 3])

#Python types will automatically be converted
import datetime
ws['A2']=datetime.datetime.now()

#save the file
wb.save("sample.xlsx")
